"""Enterprise organizations, members, and tenant knowledge base helpers."""
from __future__ import annotations

import base64
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional
from urllib.parse import quote

import requests

from kb_translate import bilingualize_kb_content
from portal_apps import default_enabled_apps, normalize_enabled_apps

KNOWLEDGE_CATEGORIES = ["光学", "机械", "材料", "热学", "电气", "控制", "其他"]
KB_CATEGORY_HEADERS = [
    "光学 / Optical",
    "机械 / Mechanical",
    "材料 / Material",
    "热学 / Thermal",
    "电气 / Electrical",
    "控制 / Control",
    "其他 / Other",
]
KB_HEADER_ROW = 3
KB_DATA_START_ROW = 4
KB_INSTRUCTION = (
    "请在下方各分类中添加经验条，每条经验占一格，经验条之间请勿隔格 / "
    "Please add your knowledge in the categories below, with one entry per cell "
    "and no empty cells in between."
)
KB_TEMPLATE_LAST_ROW = 201
KB_COLUMN_COUNT = len(KNOWLEDGE_CATEGORIES)
KB_TEMPLATE_PATH = Path(__file__).resolve().parent / "templates" / "knowledge_base_enterprise.xlsx"
KB_COLUMN_WIDTHS = {
    "A": 26.5703125,
    "B": 26.7109375,
    "C": 27.0,
    "D": 26.140625,
    "E": 25.7109375,
    "F": 26.28515625,
    "G": 27.42578125,
}
LOGO_MAX_BYTES = 500_000
LOGO_ALLOWED_TYPES = {"image/png", "image/jpeg", "image/jpg", "image/webp"}
DEFAULT_CONTRACT_YEARS = 1


def organization_display_name(
    *,
    lang: str = "zh",
    name_zh: Any = None,
    name_en: Any = None,
    name: Any = None,
    org: Optional[Dict[str, Any]] = None,
) -> str:
    if org is not None:
        name_zh = org.get("name_zh")
        name_en = org.get("name_en")
        name = org.get("name")
    zh = str(name_zh or "").strip()
    en = str(name_en or "").strip()
    legacy = str(name or "").strip()
    if lang == "en":
        return en or zh or legacy
    return zh or en or legacy


def organization_names_payload(name_zh: str, name_en: str) -> Optional[Dict[str, Any]]:
    zh = (name_zh or "").strip()
    en = (name_en or "").strip()
    if not zh and not en:
        return None
    canonical = zh or en
    return {
        "name_zh": zh or None,
        "name_en": en or None,
        "name": canonical,
    }


def contract_expires_after_years(years: int = DEFAULT_CONTRACT_YEARS) -> str:
    base = datetime.now()
    try:
        expires = base.replace(year=base.year + years)
    except ValueError:
        expires = base + timedelta(days=365 * years)
    return expires.isoformat()


def parse_contract_expires_date(value: Any) -> Optional[date]:
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        if text.endswith("Z"):
            text = text[:-1] + "+00:00"
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def format_contract_expires(value: Any, fallback: str = "-") -> str:
    parsed = parse_contract_expires_date(value)
    return parsed.isoformat() if parsed else fallback


def contract_expires_at_from_date(value: date) -> str:
    return datetime.combine(value, time(23, 59, 59)).isoformat()


def _parse_auth_users(data: Any) -> List[Dict[str, Any]]:
    if isinstance(data, dict):
        users = data.get("users")
        return users if isinstance(users, list) else []
    if isinstance(data, list):
        return data
    return []


def _table_url(supabase_url: str, table: str, query: str = "") -> str:
    base = f"{supabase_url}/rest/v1/{table}"
    return f"{base}?{query}" if query else base


def supabase_select(
    supabase_url: str,
    headers: Dict[str, str],
    table: str,
    *,
    select: str = "*",
    filters: Optional[Dict[str, str]] = None,
    order: Optional[str] = None,
) -> List[Dict[str, Any]]:
    parts = [f"select={quote(select, safe='*,()')}"]
    if filters:
        for key, value in filters.items():
            if value is None:
                parts.append(f"{key}=is.null")
            else:
                parts.append(f"{key}=eq.{quote(str(value), safe='')}")
    if order:
        parts.append(f"order={order}")
    url = _table_url(supabase_url, table, "&".join(parts))
    try:
        response = requests.get(url, headers=headers, timeout=15)
        if response.status_code == 200:
            return response.json()
    except Exception:
        pass
    return []


def supabase_insert(
    supabase_url: str,
    headers: Dict[str, str],
    table: str,
    data: Dict[str, Any],
) -> Optional[Dict[str, Any]]:
    insert_headers = {**headers, "Prefer": "return=representation"}
    try:
        response = requests.post(
            _table_url(supabase_url, table),
            headers=insert_headers,
            json=data,
            timeout=15,
        )
        if response.status_code in (200, 201):
            try:
                rows = response.json()
                if isinstance(rows, list) and rows:
                    return rows[0]
            except ValueError:
                pass
            return data
    except Exception:
        pass
    return None


def supabase_update(
    supabase_url: str,
    headers: Dict[str, str],
    table: str,
    row_id: str,
    data: Dict[str, Any],
    *,
    id_field: str = "id",
) -> tuple[bool, str]:
    url = _table_url(supabase_url, table, f"{id_field}=eq.{row_id}")
    try:
        response = requests.patch(url, headers=headers, json=data, timeout=15)
        if response.status_code in (200, 204):
            return True, ""
        return False, response.text or f"HTTP {response.status_code}"
    except Exception as exc:
        return False, str(exc)


def supabase_delete(
    supabase_url: str,
    headers: Dict[str, str],
    table: str,
    query: str,
) -> bool:
    url = _table_url(supabase_url, table, query)
    try:
        response = requests.delete(url, headers=headers, timeout=15)
        return response.status_code in (200, 204)
    except Exception:
        return False


def get_full_profile(
    supabase_url: str,
    headers: Dict[str, str],
    user_id: str,
) -> Dict[str, Any]:
    default = {
        "id": user_id,
        "email": "",
        "subscription_tier": "free",
        "free_trials_remaining": 30,
        "subscription_expires_at": None,
        "organization_id": None,
        "org_role": None,
        "organization_name": None,
        "organization_name_zh": None,
        "organization_name_en": None,
        "organization_logo_url": None,
        "max_seats": None,
    }
    if not user_id or user_id == "admin":
        return default

    rows = supabase_select(
        supabase_url,
        headers,
        "profiles",
        select="id,email,subscription_tier,free_trials_remaining,subscription_expires_at,organization_id,org_role",
        filters={"id": user_id},
    )
    if not rows:
        rows = supabase_select(
            supabase_url,
            headers,
            "profiles",
            select="id,email,subscription_tier,free_trials_remaining,subscription_expires_at",
            filters={"id": user_id},
        )
    if not rows:
        return default

    profile = {**default, **rows[0]}
    org_id = profile.get("organization_id")
    if org_id:
        org_rows = supabase_select(
            supabase_url,
            headers,
            "organizations",
            select="id,name,name_zh,name_en,max_seats,is_active,contract_expires_at,logo_url,enabled_apps",
            filters={"id": org_id},
        )
        if org_rows:
            profile["organization_name_zh"] = (org_rows[0].get("name_zh") or org_rows[0].get("name") or "").strip() or None
            profile["organization_name_en"] = (org_rows[0].get("name_en") or "").strip() or None
            profile["organization_name"] = organization_display_name(org=org_rows[0], lang="zh")
            profile["organization_logo_url"] = org_rows[0].get("logo_url")
            profile["max_seats"] = org_rows[0].get("max_seats")
            profile["org_is_active"] = org_rows[0].get("is_active", True)
            profile["enabled_apps"] = normalize_enabled_apps(org_rows[0].get("enabled_apps"))
            profile["contract_expires_at"] = org_rows[0].get("contract_expires_at")
    return profile


def is_enterprise_user(profile: Dict[str, Any]) -> bool:
    return bool(profile.get("organization_id"))


def is_org_admin(profile: Dict[str, Any]) -> bool:
    return is_enterprise_user(profile) and profile.get("org_role") == "admin"


def list_organizations(supabase_url: str, headers: Dict[str, str]) -> List[Dict[str, Any]]:
    return supabase_select(
        supabase_url,
        headers,
        "organizations",
        select="id,name,name_zh,name_en,max_seats,is_active,contract_expires_at,logo_url,enabled_apps,created_at",
        order="name.asc",
    )


def create_organization(
    supabase_url: str,
    headers: Dict[str, str],
    max_seats: int,
    *,
    name_zh: str = "",
    name_en: str = "",
    contract_years: int = DEFAULT_CONTRACT_YEARS,
) -> Optional[Dict[str, Any]]:
    names = organization_names_payload(name_zh, name_en)
    if not names:
        return None
    return supabase_insert(
        supabase_url,
        headers,
        "organizations",
        {
            **names,
            "max_seats": max_seats,
            "is_active": True,
            "enabled_apps": default_enabled_apps(),
            "contract_expires_at": contract_expires_after_years(contract_years),
            "created_at": datetime.now().isoformat(),
        },
    )


def update_organization(
    supabase_url: str,
    headers: Dict[str, str],
    org_id: str,
    data: Dict[str, Any],
) -> bool:
    ok, _ = supabase_update(supabase_url, headers, "organizations", org_id, data)
    return ok


def set_organization_names(
    supabase_url: str,
    headers: Dict[str, str],
    org_id: str,
    name_zh: str,
    name_en: str,
) -> tuple[bool, str]:
    names = organization_names_payload(name_zh, name_en)
    if not names:
        return False, "name_required"
    ok, detail = supabase_update(
        supabase_url,
        headers,
        "organizations",
        org_id,
        names,
    )
    if ok:
        return True, ""
    lowered = (detail or "").lower()
    if "name_zh" in lowered or "name_en" in lowered:
        if "column" in lowered:
            return False, "missing_column"
    return False, detail or "update_failed"


def set_organization_enabled_apps(
    supabase_url: str,
    headers: Dict[str, str],
    org_id: str,
    app_keys: List[str],
) -> tuple[bool, str]:
    enabled = normalize_enabled_apps(app_keys)
    ok, detail = supabase_update(
        supabase_url,
        headers,
        "organizations",
        org_id,
        {"enabled_apps": enabled},
    )
    if ok:
        return True, ""
    if "enabled_apps" in (detail or "") and "column" in (detail or "").lower():
        return False, "missing_column"
    return False, detail or "update_failed"


def set_organization_logo(
    supabase_url: str,
    headers: Dict[str, str],
    org_id: str,
    file_bytes: bytes,
    content_type: str,
) -> tuple[bool, str]:
    if not org_id:
        return False, "missing_org_id"
    if not file_bytes:
        return False, "empty_file"
    if len(file_bytes) > LOGO_MAX_BYTES:
        return False, "too_large"

    normalized_type = (content_type or "").lower()
    if normalized_type == "image/jpg":
        normalized_type = "image/jpeg"
    if normalized_type not in LOGO_ALLOWED_TYPES:
        return False, "invalid_type"

    encoded = base64.b64encode(file_bytes).decode("ascii")
    data_url = f"data:{normalized_type};base64,{encoded}"
    ok, detail = supabase_update(
        supabase_url,
        headers,
        "organizations",
        org_id,
        {"logo_url": data_url},
    )
    if ok:
        return True, ""
    if "logo_url" in (detail or "") and "column" in (detail or "").lower():
        return False, "missing_column"
    return False, detail or "update_failed"


def clear_organization_logo(
    supabase_url: str,
    headers: Dict[str, str],
    org_id: str,
) -> tuple[bool, str]:
    if not org_id:
        return False, "missing_org_id"
    ok, detail = supabase_update(
        supabase_url,
        headers,
        "organizations",
        org_id,
        {"logo_url": None},
    )
    if ok:
        return True, ""
    return False, detail or "update_failed"


def delete_organization(
    supabase_url: str,
    headers: Dict[str, str],
    org_id: str,
) -> tuple[bool, str]:
    try:
        if not org_id:
            return False, "missing_org_id"

        members = list_org_members(supabase_url, headers, org_id)
        for member in members:
            member_id = member.get("id")
            if member_id:
                ok, detail = supabase_update(
                    supabase_url,
                    headers,
                    "profiles",
                    member_id,
                    {"organization_id": None, "org_role": None},
                )
                if not ok:
                    return False, detail or f"unbind_failed:{member_id}"

        supabase_delete(
            supabase_url,
            headers,
            "knowledge_base",
            f"organization_id=eq.{quote(str(org_id), safe='')}&scope=eq.tenant",
        )

        if not supabase_delete(
            supabase_url,
            headers,
            "organizations",
            f"id=eq.{quote(str(org_id), safe='')}",
        ):
            return False, "delete_failed"
        return True, "ok"
    except Exception as exc:
        return False, str(exc)


def list_org_members(
    supabase_url: str,
    headers: Dict[str, str],
    organization_id: str,
) -> List[Dict[str, Any]]:
    return supabase_select(
        supabase_url,
        headers,
        "profiles",
        select="id,email,org_role,subscription_tier",
        filters={"organization_id": organization_id},
        order="email.asc",
    )


def count_org_members(
    supabase_url: str,
    headers: Dict[str, str],
    organization_id: str,
) -> int:
    return len(list_org_members(supabase_url, headers, organization_id))


def create_auth_user(
    supabase_url: str,
    service_headers: Dict[str, str],
    email: str,
    password: str,
) -> Optional[str]:
    """Create Supabase Auth user; returns user id or None."""
    try:
        response = requests.post(
            f"{supabase_url}/auth/v1/admin/users",
            headers=service_headers,
            json={"email": email, "password": password, "email_confirm": True},
            timeout=15,
        )
        if response.status_code in (200, 201):
            try:
                body = response.json()
                if isinstance(body, dict):
                    return body.get("id")
            except ValueError:
                pass
        if response.status_code == 422 and "already" in response.text.lower():
            return find_auth_user_id_by_email(supabase_url, service_headers, email)
    except Exception:
        pass
    return None


def ensure_profile(
    supabase_url: str,
    headers: Dict[str, str],
    user_id: str,
    email: str,
    data: Dict[str, Any],
) -> tuple[bool, str]:
    rows = supabase_select(
        supabase_url,
        headers,
        "profiles",
        filters={"id": user_id},
    )
    payload = {"email": email, **data}
    if rows:
        return supabase_update(supabase_url, headers, "profiles", user_id, payload)
    payload["id"] = user_id
    payload.setdefault("free_trials_remaining", 0)
    payload.setdefault("subscription_tier", "pro")
    result = supabase_insert(supabase_url, headers, "profiles", payload)
    if result is not None:
        return True, ""
    return False, "insert_failed"


def find_auth_user_id_by_email(
    supabase_url: str,
    service_headers: Dict[str, str],
    email: str,
) -> Optional[str]:
    normalized = (email or "").strip().lower()
    if not normalized:
        return None

    page = 1
    per_page = 200
    while page <= 20:
        try:
            response = requests.get(
                f"{supabase_url}/auth/v1/admin/users",
                headers=service_headers,
                params={"page": page, "per_page": per_page},
                timeout=15,
            )
            if response.status_code != 200:
                break
            users = _parse_auth_users(response.json())
            if not users:
                break
            for user in users:
                if (user.get("email") or "").strip().lower() == normalized:
                    return user.get("id")
            if len(users) < per_page:
                break
            page += 1
        except Exception:
            break
    return None


def find_user_id_by_email(
    supabase_url: str,
    service_headers: Dict[str, str],
    email: str,
) -> Optional[str]:
    user_id = find_auth_user_id_by_email(supabase_url, service_headers, email)
    if user_id:
        return user_id

    rows = supabase_select(
        supabase_url,
        service_headers,
        "profiles",
        select="id,email",
        filters={"email": (email or "").strip()},
    )
    if rows:
        return rows[0].get("id")

    return None


def set_auth_user_password(
    supabase_url: str,
    service_headers: Dict[str, str],
    user_id: str,
    password: str,
) -> tuple[bool, str]:
    try:
        response = requests.put(
            f"{supabase_url}/auth/v1/admin/users/{user_id}",
            headers=service_headers,
            json={"password": password, "email_confirm": True},
            timeout=15,
        )
        if response.status_code in (200, 201):
            return True, ""
        return False, response.text or f"HTTP {response.status_code}"
    except Exception as exc:
        return False, str(exc)


def verify_login_credentials(
    supabase_url: str,
    anon_key: str,
    email: str,
    password: str,
) -> tuple[bool, str]:
    headers = {
        "apikey": anon_key,
        "Content-Type": "application/json",
    }
    try:
        response = requests.post(
            f"{supabase_url}/auth/v1/token?grant_type=password",
            headers=headers,
            json={"email": (email or "").strip(), "password": password},
            timeout=15,
        )
        if response.status_code == 200:
            return True, ""
        try:
            body = response.json()
            return False, (
                body.get("error_description")
                or body.get("msg")
                or body.get("message")
                or response.text
            )
        except ValueError:
            return False, response.text or f"HTTP {response.status_code}"
    except Exception as exc:
        return False, str(exc)


def change_user_password(
    supabase_url: str,
    anon_key: str,
    email: str,
    current_password: str,
    new_password: str,
) -> tuple[bool, str]:
    headers = {
        "apikey": anon_key,
        "Content-Type": "application/json",
    }
    try:
        login_response = requests.post(
            f"{supabase_url}/auth/v1/token?grant_type=password",
            headers=headers,
            json={"email": (email or "").strip(), "password": current_password},
            timeout=15,
        )
        if login_response.status_code != 200:
            return False, "current_password_wrong"

        access_token = login_response.json().get("access_token")
        if not access_token:
            return False, "current_password_wrong"

        update_response = requests.put(
            f"{supabase_url}/auth/v1/user",
            headers={
                **headers,
                "Authorization": f"Bearer {access_token}",
            },
            json={"password": new_password},
            timeout=15,
        )
        if update_response.status_code in (200, 201):
            return True, ""
        try:
            body = update_response.json()
            return False, (
                body.get("error_description")
                or body.get("msg")
                or body.get("message")
                or update_response.text
            )
        except ValueError:
            return False, update_response.text or f"HTTP {update_response.status_code}"
    except Exception as exc:
        return False, str(exc)


def assign_or_provision_org_user(
    supabase_url: str,
    service_headers: Dict[str, str],
    email: str,
    password: str,
    organization_id: str,
    org_role: str,
    *,
    max_seats: int = 500,
) -> tuple[bool, str]:
    """Create or locate auth user, always set password, then bind to organization."""
    normalized_email = (email or "").strip()
    if not normalized_email:
        return False, "email_required"
    if not password or len(password) < 6:
        return False, "password_required"

    user_id = find_auth_user_id_by_email(supabase_url, service_headers, normalized_email)
    if not user_id:
        if count_org_members(supabase_url, service_headers, organization_id) >= max_seats:
            return False, "seat_limit"
        user_id = create_auth_user(supabase_url, service_headers, normalized_email, password)
        if not user_id:
            return False, "auth_failed"

    pwd_ok, pwd_detail = set_auth_user_password(
        supabase_url, service_headers, user_id, password
    )
    if not pwd_ok:
        return False, pwd_detail or "password_reset_failed"

    ok, detail = ensure_profile(
        supabase_url,
        service_headers,
        user_id,
        normalized_email,
        {
            "organization_id": organization_id,
            "org_role": org_role,
            "subscription_tier": "pro",
            "free_trials_remaining": 0,
        },
    )
    return (ok, "ok" if ok else (detail or "profile_failed"))


def assign_email_to_org(
    supabase_url: str,
    service_headers: Dict[str, str],
    email: str,
    organization_id: str,
    org_role: str,
) -> tuple[bool, str]:
    user_id = find_user_id_by_email(supabase_url, service_headers, email)
    if not user_id:
        return False, "not_found"

    ok, detail = ensure_profile(
        supabase_url,
        service_headers,
        user_id,
        email.strip(),
        {
            "organization_id": organization_id,
            "org_role": org_role,
            "subscription_tier": "pro",
        },
    )
    if not ok:
        return False, detail or "profile_failed"
    return True, "ok"


def assign_user_to_org(
    supabase_url: str,
    headers: Dict[str, str],
    user_id: str,
    organization_id: Optional[str],
    org_role: Optional[str],
    *,
    make_enterprise: bool = True,
) -> bool:
    data: Dict[str, Any] = {
        "organization_id": organization_id,
        "org_role": org_role,
    }
    if make_enterprise and organization_id:
        data["subscription_tier"] = "pro"
    elif not organization_id:
        data["org_role"] = None
    ok, _ = supabase_update(supabase_url, headers, "profiles", user_id, data)
    return ok


def add_org_member(
    supabase_url: str,
    service_headers: Dict[str, str],
    organization_id: str,
    email: str,
    password: str,
    org_role: str = "member",
    max_seats: int = 10,
) -> tuple[bool, str]:
    if count_org_members(supabase_url, service_headers, organization_id) >= max_seats:
        return False, "seat_limit"

    normalized_email = (email or "").strip()
    user_id = find_auth_user_id_by_email(supabase_url, service_headers, normalized_email)
    if not user_id:
        user_id = create_auth_user(supabase_url, service_headers, normalized_email, password)
    if not user_id:
        return False, "auth_failed"

    pwd_ok, pwd_detail = set_auth_user_password(
        supabase_url, service_headers, user_id, password
    )
    if not pwd_ok:
        return False, pwd_detail or "password_reset_failed"

    ok, reason = ensure_profile(
        supabase_url,
        service_headers,
        user_id,
        normalized_email,
        {
            "organization_id": organization_id,
            "org_role": org_role,
            "subscription_tier": "pro",
            "free_trials_remaining": 0,
        },
    )
    return (ok, "ok" if ok else (reason or "profile_failed"))


def remove_org_member(
    supabase_url: str,
    headers: Dict[str, str],
    user_id: str,
) -> bool:
    return assign_user_to_org(
        supabase_url,
        headers,
        user_id,
        organization_id=None,
        org_role=None,
        make_enterprise=False,
    )


def list_tenant_knowledge(
    supabase_url: str,
    headers: Dict[str, str],
    organization_id: str,
) -> List[Dict[str, Any]]:
    return supabase_select(
        supabase_url,
        headers,
        "knowledge_base",
        select="id,category,content,content_en,created_at",
        filters={"scope": "tenant", "organization_id": organization_id},
        order="id.desc",
    )


def add_tenant_knowledge(
    supabase_url: str,
    headers: Dict[str, str],
    organization_id: str,
    category: str,
    content: str,
    *,
    content_en: Optional[str] = None,
) -> bool:
    result = supabase_insert(
        supabase_url,
        headers,
        "knowledge_base",
        {
            "category": category,
            "content": content,
            "content_en": content_en or content,
            "scope": "tenant",
            "organization_id": organization_id,
            "created_at": datetime.now().isoformat(),
        },
    )
    return result is not None


def delete_tenant_knowledge(
    supabase_url: str,
    headers: Dict[str, str],
    record_id: int,
    organization_id: str,
) -> bool:
    rows = supabase_select(
        supabase_url,
        headers,
        "knowledge_base",
        filters={"id": str(record_id), "organization_id": organization_id, "scope": "tenant"},
    )
    if not rows:
        return False
    return supabase_delete(
        supabase_url,
        headers,
        "knowledge_base",
        f"id=eq.{record_id}&organization_id=eq.{organization_id}&scope=eq.tenant",
    )


def _kb_workbook_title(org_name: str, lang: str) -> str:
    if lang == "en":
        return f"{org_name} · Enterprise Knowledge Database"
    return f"{org_name} · 企业数据库"


def _open_kb_template_workbook():
    from openpyxl import load_workbook

    if not KB_TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Missing knowledge base template: {KB_TEMPLATE_PATH}")
    return load_workbook(KB_TEMPLATE_PATH)


def _save_workbook_bytes(workbook) -> bytes:
    import io

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _apply_kb_title(worksheet, org_name: str, lang: str) -> None:
    from openpyxl.styles import Font

    title_cell = worksheet.cell(1, 1)
    title_cell.value = _kb_workbook_title(org_name, lang)
    title_cell.font = Font(name="Calibri", size=16, bold=True)
    worksheet.row_dimensions[1].height = 21.0

    instruction_cell = worksheet.cell(2, 1)
    instruction_cell.value = KB_INSTRUCTION
    instruction_cell.font = Font(name="Calibri", size=10, bold=False)


def _ensure_kb_layout(worksheet) -> None:
    from openpyxl.styles import Alignment, Border, Font, Side

    thin = Side(style="thin")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Calibri", size=11, bold=True)
    header_alignment = Alignment(horizontal="center")

    for column_letter, width in KB_COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column_letter].width = width

    for col_idx, header in enumerate(KB_CATEGORY_HEADERS, start=1):
        header_cell = worksheet.cell(KB_HEADER_ROW, col_idx)
        header_cell.value = header
        header_cell.font = header_font
        header_cell.border = header_border
        header_cell.alignment = header_alignment

    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row_idx in range(KB_DATA_START_ROW, KB_TEMPLATE_LAST_ROW + 1):
        for col_idx in range(1, KB_COLUMN_COUNT + 1):
            data_cell = worksheet.cell(row_idx, col_idx)
            data_cell.border = data_border


def _clear_kb_data_area(worksheet) -> None:
    for row_idx in range(KB_DATA_START_ROW, KB_TEMPLATE_LAST_ROW + 1):
        for col_idx in range(1, KB_COLUMN_COUNT + 1):
            worksheet.cell(row_idx, col_idx).value = None


def _fill_kb_data_area(worksheet, entries_by_category: Dict[str, List[str]]) -> None:
    max_rows = max((len(values) for values in entries_by_category.values()), default=0)
    for row_offset in range(max_rows):
        row_idx = KB_DATA_START_ROW + row_offset
        for col_idx, category in enumerate(KNOWLEDGE_CATEGORIES, start=1):
            values = entries_by_category.get(category, [])
            if row_offset < len(values):
                worksheet.cell(row_idx, col_idx, values[row_offset])


def _prepare_kb_workbook(org_name: str, lang: str):
    workbook = _open_kb_template_workbook()
    worksheet = workbook.active
    _apply_kb_title(worksheet, org_name, lang)
    _ensure_kb_layout(worksheet)
    _clear_kb_data_area(worksheet)
    return workbook, worksheet


def _parse_category_label(label: Any) -> Optional[str]:
    if label is None:
        return None
    text = str(label).strip()
    if not text or text.lower() == "nan":
        return None

    if text in KNOWLEDGE_CATEGORIES:
        return text

    for category, header in zip(KNOWLEDGE_CATEGORIES, KB_CATEGORY_HEADERS):
        if text == header:
            return category
        if text.lower() == header.lower():
            return category

    if "/" in text:
        zh_part = text.split("/", 1)[0].strip()
        if zh_part in KNOWLEDGE_CATEGORIES:
            return zh_part
        en_part = text.split("/", 1)[1].strip().lower()
        en_map = {
            "optical": "光学",
            "mechanical": "机械",
            "material": "材料",
            "thermal": "热学",
            "electrical": "电气",
            "control": "控制",
            "other": "其他",
        }
        if en_part in en_map:
            return en_map[en_part]

    lowered = text.lower()
    if lowered == "other":
        return "其他"
    return None


def build_kb_template_excel(org_name: str, lang: str = "zh") -> bytes:
    workbook, _worksheet = _prepare_kb_workbook(org_name, lang)
    return _save_workbook_bytes(workbook)


def build_kb_export_excel(
    entries: List[Dict[str, Any]],
    org_name: str,
    lang: str = "zh",
) -> bytes:
    grouped: Dict[str, List[str]] = {category: [] for category in KNOWLEDGE_CATEGORIES}
    for entry in entries:
        category = _parse_category_label(entry.get("category")) or str(entry.get("category", "")).strip()
        content = str(entry.get("content", "")).strip()
        if not category or not content:
            continue
        if category not in grouped:
            grouped[category] = []
        grouped[category].append(content)

    workbook, worksheet = _prepare_kb_workbook(org_name, lang)
    _fill_kb_data_area(worksheet, grouped)
    return _save_workbook_bytes(workbook)


def _load_kb_worksheet(file_bytes: bytes):
    import io

    from openpyxl import load_workbook

    return load_workbook(io.BytesIO(file_bytes)).active


def _map_kb_header_columns(worksheet, header_row: int) -> Dict[int, str]:
    categories_by_col: Dict[int, str] = {}
    for col_idx in range(1, KB_COLUMN_COUNT + 1):
        category = _parse_category_label(worksheet.cell(header_row, col_idx).value)
        if category:
            categories_by_col[col_idx] = category
    return categories_by_col


def _resolve_kb_header_row(worksheet) -> Optional[int]:
    """Enterprise KB import always expects bilingual headers on row 3, cols A-G."""
    row_three = _map_kb_header_columns(worksheet, KB_HEADER_ROW)
    if (
        len(row_three) == KB_COLUMN_COUNT
        and set(row_three.values()) == set(KNOWLEDGE_CATEGORIES)
        and row_three.get(KB_COLUMN_COUNT) == "其他"
    ):
        return KB_HEADER_ROW

    for row_idx in range(1, 8):
        if row_idx == KB_HEADER_ROW:
            continue
        mapped = _map_kb_header_columns(worksheet, row_idx)
        if len(mapped) == KB_COLUMN_COUNT and "其他" in mapped.values():
            return row_idx
    return None


def _parse_wide_kb_worksheet(worksheet) -> Optional[List[Dict[str, str]]]:
    """
    Parse enterprise wide-format workbook.
    - Row 3: category headers including「其他 / Other」(column G)
    - Row 4+: one knowledge item per cell, grouped by column
    """
    header_row = _resolve_kb_header_row(worksheet)
    if not header_row:
        return None

    categories_by_col = _map_kb_header_columns(worksheet, header_row)
    if "其他" not in categories_by_col.values():
        return None

    rows_to_import: List[Dict[str, str]] = []
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        for col_idx, category in categories_by_col.items():
            raw_value = worksheet.cell(row_idx, col_idx).value
            if raw_value is None:
                continue
            content = str(raw_value).strip()
            if not content or content.lower() == "nan":
                continue
            rows_to_import.append(
                {
                    "category": category,
                    "content": content,
                    "content_en": content,
                }
            )
    return rows_to_import


def _parse_legacy_kb_dataframe(file_bytes: bytes) -> Optional[List[Dict[str, str]]]:
    import io

    import pandas as pd

    legacy_aliases = {
        "category": {"分类", "category", "Category"},
        "content": {"经验内容", "content", "Content"},
        "content_en": {"经验内容(英文)", "content_en", "Content (EN)", "Content_EN"},
    }

    try:
        frame = pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl")
    except Exception:
        return None

    if frame.empty:
        return None

    column_map: Dict[str, str] = {}
    for column in frame.columns:
        cleaned = str(column).strip()
        for field, aliases in legacy_aliases.items():
            if cleaned in aliases:
                column_map[field] = column

    if "category" not in column_map or "content" not in column_map:
        return None

    rows_to_import: List[Dict[str, str]] = []
    for _, row in frame.iterrows():
        category = _parse_category_label(row.get(column_map["category"])) or str(
            row.get(column_map["category"], "")
        ).strip()
        content = str(row.get(column_map["content"], "")).strip()
        if not content:
            continue
        content_en_col = column_map.get("content_en")
        content_en = str(row.get(content_en_col, "")).strip() if content_en_col else content
        if not content_en:
            content_en = content
        rows_to_import.append(
            {
                "category": category or KNOWLEDGE_CATEGORIES[0],
                "content": content,
                "content_en": content_en,
            }
        )
    return rows_to_import or None


def delete_all_tenant_knowledge(
    supabase_url: str,
    headers: Dict[str, str],
    organization_id: str,
) -> bool:
    if not organization_id:
        return False
    return supabase_delete(
        supabase_url,
        headers,
        "knowledge_base",
        f"organization_id=eq.{quote(str(organization_id), safe='')}&scope=eq.tenant",
    )


def import_tenant_knowledge_excel(
    supabase_url: str,
    headers: Dict[str, str],
    organization_id: str,
    file_bytes: bytes,
    *,
    replace_existing: bool = False,
    translate_to_en: Optional[Callable[[str], str]] = None,
    translate_to_zh: Optional[Callable[[str], str]] = None,
) -> tuple[int, str]:
    if not organization_id:
        return 0, "missing_org_id"
    if not file_bytes:
        return 0, "empty_file"

    rows_to_import: Optional[List[Dict[str, str]]] = None
    try:
        worksheet = _load_kb_worksheet(file_bytes)
        rows_to_import = _parse_wide_kb_worksheet(worksheet)
    except Exception as exc:
        return 0, f"invalid_excel:{exc}"

    if rows_to_import is None:
        rows_to_import = _parse_legacy_kb_dataframe(file_bytes)

    if rows_to_import is None:
        return 0, "missing_columns"

    if not rows_to_import:
        return 0, "no_valid_rows"

    if replace_existing:
        delete_all_tenant_knowledge(supabase_url, headers, organization_id)

    imported = 0
    for item in rows_to_import:
        zh_text, en_text = bilingualize_kb_content(
            item["content"],
            translate_to_en=translate_to_en,
            translate_to_zh=translate_to_zh,
        )
        if add_tenant_knowledge(
            supabase_url,
            headers,
            organization_id,
            item["category"],
            zh_text,
            content_en=en_text,
        ):
            imported += 1

    if imported == 0:
        return 0, "import_failed"
    return imported, "ok"
